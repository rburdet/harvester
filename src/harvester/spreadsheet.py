from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .aggregate import DayEntry


HEADERS = ["Date", "Hours", "Tickets", "Carried", "Commits"]


def _rows(entries: list[DayEntry]) -> list[list]:
    return [
        [
            e.day.isoformat(),
            round(e.hours, 2),
            ", ".join(e.tickets),
            e.carried_from.isoformat() if e.carried_from else "",
            e.commit_count,
        ]
        for e in entries
    ]


def write(entries: list[DayEntry], out_dir: Path, month_label: str) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / f"timesheet_{month_label}.xlsx"
    csv_path = out_dir / f"timesheet_{month_label}.csv"
    rows = _rows(entries)
    total = round(sum(e.hours for e in entries), 2)

    wb = Workbook()
    ws = wb.active
    ws.title = month_label
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    ws.append([])
    ws.append(["Total", total, "", "", ""])
    ws[ws.max_row][0].font = Font(bold=True)
    ws[ws.max_row][1].font = Font(bold=True)
    for col_idx, width in enumerate([12, 8, 40, 12, 8], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    wb.save(xlsx_path)

    with csv_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(rows)
        w.writerow([])
        w.writerow(["Total", total, "", "", ""])

    return xlsx_path, csv_path
